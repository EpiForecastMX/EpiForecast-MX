# scripts/compara_metricas.py
"""Genera un Excel comparativo de metricas Prophet vs DeepAR.

Uso:
    python -m scripts.compara_metricas
    make compare-metrics
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from epiforecast.utils.config import logger

# ---------------------------------------------------------------------------
# Constantes de estilo (colores del proyecto)
# ---------------------------------------------------------------------------
TEAL = "004D40"
BURGUNDY = "880E4F"
AMBER = "FF6F00"
DARK_GRAY = "424242"
GREEN_BG = "C8E6C9"
RED_BG = "FFCDD2"
WHITE_FG = "FFFFFF"

METRIC_COLS = ["rmse", "mae", "mape", "mase"]
PROPHET_HYPERPARAMS = ["seasonality_mode", "changepoint_prior_scale", "seasonality_prior_scale"]
DEEPAR_HYPERPARAMS = [
    "epochs",
    "context_length",
    "num_layers",
    "learning_rate",
    "scaling",
    "distr_output",
]

MODELS_DIR = Path("models")
OUTPUT_DIR = Path("reports/forecasts/comparacion_modelos")
OUTPUT_FILE = OUTPUT_DIR / "comparacion_metricas.xlsx"
TZ_CDMX = ZoneInfo("America/Mexico_City")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Mapeo de columnas internas -> nombres legibles en el Excel
_COL_DISPLAY: dict[str, str] = {
    "padecimiento": "Padecimiento",
    "sexo": "Sexo",
    "nivel": "Nivel",
    "Entidad": "Entidad",
    "rmse_prophet": "RMSE Prophet",
    "rmse_deepar": "RMSE DeepAR",
    "delta_rmse": "Delta RMSE",
    "pct_rmse": "% Diferencia RMSE",
    "mae_prophet": "MAE Prophet",
    "mae_deepar": "MAE DeepAR",
    "delta_mae": "Delta MAE",
    "pct_mae": "% Diferencia MAE",
    "mape_prophet": "MAPE Prophet",
    "mape_deepar": "MAPE DeepAR",
    "delta_mape": "Delta MAPE",
    "pct_mape": "% Diferencia MAPE",
    "mase_prophet": "MASE Prophet",
    "mase_deepar": "MASE DeepAR",
    "delta_mase": "Delta MASE",
    "pct_mase": "% Diferencia MASE",
    "ganador_rmse": "Ganador RMSE",
    "ganador_mae": "Ganador MAE",
    "ganador_mape": "Ganador MAPE",
    "ganador_mase": "Ganador MASE",
    "confianza_prophet": "Confianza Prophet",
    "confianza_deepar": "Confianza DeepAR",
    "promedio_semanal_prophet": "Prom. Semanal Prophet",
    "promedio_semanal_deepar": "Prom. Semanal DeepAR",
    "tiempo_total_seg_prophet": "Tiempo (s) Prophet",
    "tiempo_total_seg_deepar": "Tiempo (s) DeepAR",
    "seasonality_mode": "Estacionalidad (Prophet)",
    "changepoint_prior_scale": "Changepoint Prior (Prophet)",
    "seasonality_prior_scale": "Seasonality Prior (Prophet)",
    "epochs": "Epochs (DeepAR)",
    "context_length": "Context Length (DeepAR)",
    "num_layers": "Num Layers (DeepAR)",
    "learning_rate": "Learning Rate (DeepAR)",
    "scaling": "Scaling (DeepAR)",
    "distr_output": "Distribucion (DeepAR)",
}


# ---------------------------------------------------------------------------
# Lectura de datos
# ---------------------------------------------------------------------------
def _leer_completos(modelo: str) -> pd.DataFrame:
    """Lee todos los CSV ``*_completo.csv`` de un modelo y los concatena."""
    base = MODELS_DIR / modelo
    archivos = sorted(base.rglob("*_completo.csv"))

    if not archivos:
        logger.warning(f"No se encontraron archivos *_completo.csv en {base}")
        return pd.DataFrame()

    frames: list[pd.DataFrame] = []
    for f in archivos:
        logger.info(f"  Leyendo {f.relative_to('.')}")
        frames.append(pd.read_csv(f))

    return pd.concat(frames, ignore_index=True)


def _normalizar_keys(df: pd.DataFrame) -> pd.DataFrame:
    """Agrega columna ``Entidad`` si no existe y rellena NaN."""
    if "Entidad" not in df.columns:
        df["Entidad"] = ""
    df["Entidad"] = df["Entidad"].fillna("")
    return df


# ---------------------------------------------------------------------------
# Construccion de hojas
# ---------------------------------------------------------------------------
def _build_detalle(prophet: pd.DataFrame, deepar: pd.DataFrame) -> pd.DataFrame:
    """Merge completo lado a lado con deltas por metrica."""
    p = _normalizar_keys(prophet.copy())
    d = _normalizar_keys(deepar.copy())

    merge_keys = ["padecimiento", "sexo", "nivel", "Entidad"]

    def _pick(df: pd.DataFrame, hyperparams: list[str]) -> pd.DataFrame:
        cols = merge_keys.copy()
        cols += [c for c in METRIC_COLS if c in df.columns]
        cols += [
            c for c in ("confianza", "promedio_semanal", "tiempo_total_seg") if c in df.columns
        ]
        cols += [c for c in hyperparams if c in df.columns]
        return df[[c for c in cols if c in df.columns]]

    pm = _pick(p, PROPHET_HYPERPARAMS)
    dm = _pick(d, DEEPAR_HYPERPARAMS)

    merged = pm.merge(dm, on=merge_keys, how="outer", suffixes=("_prophet", "_deepar"))

    # Calcular deltas y % de mejora por metrica
    for m in METRIC_COLS:
        pc, dc = f"{m}_prophet", f"{m}_deepar"
        if pc in merged.columns and dc in merged.columns:
            merged[f"delta_{m}"] = merged[dc] - merged[pc]
            safe_denom = merged[pc].replace(0, float("nan"))
            merged[f"pct_{m}"] = ((merged[dc] - merged[pc]) / safe_denom) * 100

    # Ordenar columnas: keys → (prophet, deepar, delta, pct) por metrica → info → hyperparams
    ordered: list[str] = list(merge_keys)
    for m in METRIC_COLS:
        for sfx in (f"{m}_prophet", f"{m}_deepar", f"delta_{m}", f"pct_{m}"):
            if sfx in merged.columns:
                ordered.append(sfx)

    info_cols = [
        "confianza_prophet",
        "confianza_deepar",
        "promedio_semanal_prophet",
        "promedio_semanal_deepar",
        "tiempo_total_seg_prophet",
        "tiempo_total_seg_deepar",
    ]
    ordered += [c for c in info_cols if c in merged.columns]
    ordered += [c for c in PROPHET_HYPERPARAMS if c in merged.columns]
    ordered += [c for c in DEEPAR_HYPERPARAMS if c in merged.columns]

    # Columnas residuales
    ordered += [c for c in merged.columns if c not in ordered]

    return merged[ordered].sort_values(merge_keys, ignore_index=True)


def _build_resumen(detalle: pd.DataFrame) -> pd.DataFrame:
    """Tabla resumen: promedio de metricas por padecimiento y modelo."""
    rows: list[dict[str, object]] = []

    for pad, grp in detalle.groupby("padecimiento"):
        row: dict[str, object] = {"padecimiento": pad}
        for m in METRIC_COLS:
            pc, dc = f"{m}_prophet", f"{m}_deepar"
            p_val = grp[pc].mean(skipna=True) if pc in grp.columns else float("nan")
            d_val = grp[dc].mean(skipna=True) if dc in grp.columns else float("nan")

            row[pc] = p_val
            row[dc] = d_val

            if pd.notna(p_val) and pd.notna(d_val):
                row[f"delta_{m}"] = d_val - p_val
                if d_val < p_val:
                    row[f"ganador_{m}"] = "DeepAR"
                elif p_val < d_val:
                    row[f"ganador_{m}"] = "Prophet"
                else:
                    row[f"ganador_{m}"] = "Empate"
            else:
                row[f"delta_{m}"] = None
                row[f"ganador_{m}"] = "N/D"

        rows.append(row)

    ordered_cols = ["padecimiento"]
    for m in METRIC_COLS:
        ordered_cols += [f"{m}_prophet", f"{m}_deepar", f"delta_{m}", f"ganador_{m}"]

    return pd.DataFrame(rows)[[c for c in ordered_cols if c in pd.DataFrame(rows).columns]]


def _build_metadata(prophet_files: list[Path], deepar_files: list[Path]) -> pd.DataFrame:
    """Hoja de metadata con informacion de generacion."""
    ahora = datetime.now(tz=TZ_CDMX)
    return pd.DataFrame(
        {
            "Campo": [
                "Fecha de generacion",
                "Zona horaria",
                "Version del proyecto",
                "Archivos Prophet",
                "Archivos DeepAR",
                "Directorio de salida",
            ],
            "Valor": [
                ahora.strftime("%Y-%m-%d %H:%M:%S"),
                "America/Mexico_City (UTC-6)",
                "2.0.0",
                ", ".join(str(f.relative_to(".")) for f in prophet_files),
                ", ".join(str(f.relative_to(".")) for f in deepar_files),
                str(OUTPUT_FILE),
            ],
        }
    )


# ---------------------------------------------------------------------------
# Estilos openpyxl
# ---------------------------------------------------------------------------
def _autofit_columns(ws) -> None:  # type: ignore[no-untyped-def]
    """Ajusta ancho de columnas al contenido."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 35)


def _aplicar_bordes(ws, max_row: int, max_col: int) -> None:  # type: ignore[no-untyped-def]
    """Bordes finos para toda la tabla."""
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER


def _aplicar_formato_numerico(ws, df: pd.DataFrame, start_row: int = 2) -> None:  # type: ignore[no-untyped-def]
    """Aplica formato numerico segun tipo de columna (usa nombres display)."""
    for col_idx, col_name in enumerate(df.columns, 1):
        col_lower = col_name.lower()

        if "% diferencia" in col_lower:
            fmt = "0.00"
        elif any(k in col_lower for k in ("rmse", "mae", "mase")) and "%" not in col_lower:
            fmt = "0.0000"
        elif "mape" in col_lower:
            fmt = "0.00"
        elif "tiempo" in col_lower:
            fmt = "0.1"
        elif "prom." in col_lower:
            fmt = "0.00"
        else:
            continue

        for row in range(start_row, start_row + len(df)):
            ws.cell(row=row, column=col_idx).number_format = fmt


def _aplicar_condicional_detalle(ws, df: pd.DataFrame, start_row: int = 2) -> None:  # type: ignore[no-untyped-def]
    """Verde si MASE < 1, rojo si confianza = 'insuficiente'."""
    green_fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
    red_fill = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")

    for col_idx, col_name in enumerate(df.columns, 1):
        col_lower = col_name.lower()
        is_mase = "mase" in col_lower and "delta" not in col_lower and "%" not in col_lower
        is_confianza = "confianza" in col_lower

        if not is_mase and not is_confianza:
            continue

        for row_idx, val in enumerate(df[col_name], start_row):
            cell = ws.cell(row=row_idx, column=col_idx)
            if is_mase and pd.notna(val) and isinstance(val, int | float) and val < 1:
                cell.fill = green_fill
            if is_confianza and val == "insuficiente":
                cell.fill = red_fill


def _header_color(col_name: str) -> str:
    """Devuelve color hex para el header segun el nombre de la columna."""
    col_lower = col_name.lower()
    if "prophet" in col_lower:
        return TEAL
    if "deepar" in col_lower:
        return BURGUNDY
    if "delta" in col_lower or "% diferencia" in col_lower:
        return AMBER
    if "ganador" in col_lower:
        return "37474F"
    return DARK_GRAY


def _estilizar_headers(ws, df: pd.DataFrame) -> None:  # type: ignore[no-untyped-def]
    """Aplica estilos de header (fondo, fuente, alineacion) segun modelo."""
    for col_idx, col_name in enumerate(df.columns, 1):
        color = _header_color(col_name)
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(bold=True, color=WHITE_FG, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _colorear_ganadores(ws, df: pd.DataFrame) -> None:  # type: ignore[no-untyped-def]
    """Colorea celdas de ganador con el color del modelo correspondiente."""
    for col_idx, col_name in enumerate(df.columns, 1):
        if "ganador" not in col_name.lower():
            continue
        for row_idx, val in enumerate(df[col_name], 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if val == "Prophet":
                cell.font = Font(bold=True, color=TEAL)
            elif val == "DeepAR":
                cell.font = Font(bold=True, color=BURGUNDY)


def _estilo_metadata(ws) -> None:  # type: ignore[no-untyped-def]
    """Estilo sobrio para hoja Metadata."""
    fill = PatternFill(start_color=DARK_GRAY, end_color=DARK_GRAY, fill_type="solid")
    font = Font(bold=True, color=WHITE_FG, size=11)
    for col in range(1, 3):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


# ---------------------------------------------------------------------------
# Generacion del Excel
# ---------------------------------------------------------------------------
def _renombrar(df: pd.DataFrame) -> pd.DataFrame:
    """Renombra columnas internas a nombres legibles."""
    return df.rename(columns=_COL_DISPLAY)


def generar_excel_comparativo() -> Path:
    """Genera el Excel comparativo y devuelve la ruta del archivo."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # --- Leer datos ---
    logger.info("Leyendo metricas Prophet...")
    prophet = _leer_completos("prophet")
    logger.info("Leyendo metricas DeepAR...")
    deepar = _leer_completos("deepar")

    if prophet.empty and deepar.empty:
        logger.error("No se encontraron datos de ningun modelo.")
        raise SystemExit(1)

    prophet_files = sorted((MODELS_DIR / "prophet").rglob("*_completo.csv"))
    deepar_files = sorted((MODELS_DIR / "deepar").rglob("*_completo.csv"))

    # --- Construir hojas ---
    logger.info("Construyendo hoja Detalle...")
    detalle = _build_detalle(prophet, deepar)

    logger.info("Construyendo hoja Resumen...")
    resumen = _build_resumen(detalle)

    logger.info("Construyendo hoja Metadata...")
    metadata = _build_metadata(prophet_files, deepar_files)

    # Versiones con nombres legibles para el Excel
    resumen_xl = _renombrar(resumen)
    detalle_xl = _renombrar(detalle)

    # --- Escribir Excel ---
    logger.info(f"Escribiendo Excel en {OUTPUT_FILE}...")
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        resumen_xl.to_excel(writer, sheet_name="Resumen", index=False)
        detalle_xl.to_excel(writer, sheet_name="Detalle", index=False)
        metadata.to_excel(writer, sheet_name="Metadata", index=False)

        wb = writer.book

        # --- Estilizar Resumen ---
        ws_res = wb["Resumen"]
        _estilizar_headers(ws_res, resumen_xl)
        _aplicar_formato_numerico(ws_res, resumen_xl)
        _colorear_ganadores(ws_res, resumen_xl)
        _aplicar_bordes(ws_res, len(resumen_xl) + 1, len(resumen_xl.columns))
        _autofit_columns(ws_res)
        ws_res.freeze_panes = "B2"

        # --- Estilizar Detalle ---
        ws_det = wb["Detalle"]
        _estilizar_headers(ws_det, detalle_xl)
        _aplicar_formato_numerico(ws_det, detalle_xl)
        _aplicar_condicional_detalle(ws_det, detalle_xl, start_row=2)
        _aplicar_bordes(ws_det, len(detalle_xl) + 1, len(detalle_xl.columns))
        _autofit_columns(ws_det)
        ws_det.freeze_panes = "E2"

        # --- Estilizar Metadata ---
        ws_meta = wb["Metadata"]
        _estilo_metadata(ws_meta)
        _aplicar_bordes(ws_meta, len(metadata) + 1, 2)
        _autofit_columns(ws_meta)

    logger.success(f"Excel generado: {OUTPUT_FILE}")
    return OUTPUT_FILE


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main() -> None:
    logger.info("Iniciando comparacion de metricas Prophet vs DeepAR...")
    generar_excel_comparativo()
    logger.success("Proceso de comparacion de metricas finalizado.")


if __name__ == "__main__":
    main()
