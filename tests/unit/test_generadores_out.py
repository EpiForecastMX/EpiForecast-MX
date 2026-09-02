"""`--out` en los tres generadores que escribían en el árbol real.

Sin `--out`, `build_web_knowledge.py`, `build_tableau.py` y `genera_validacion_semanal.py`
escribían en rutas fijas del repositorio, y el refresh semanal ensuciaba el árbol real en
cada corrida. Con `--out` el refresh sellado los apunta al staging; sin él conservan la
ruta legacy, así que nada cambia para quien los usa a mano.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
import pandas as pd


def test_build_web_knowledge_escribe_donde_se_le_dice(tmp_path: Path) -> None:
    from scripts import build_web_knowledge as g

    assert g._parse_args([]).out == g.OUTPUT
    destino = tmp_path / "staging" / "epibot" / "knowledge.json"
    assert g._parse_args(["--out", str(destino)]).out == destino

    g.escribe_knowledge({"a": float("nan"), "b": [1, float("nan")], "c": {"d": "ñ"}}, destino)

    # NaN -> null también anidado: el encoder anterior nunca lo convertía y podía emitir
    # el token `NaN`, que no es JSON y rompe `JSON.parse` en el navegador.
    assert destino.read_text(encoding="utf-8") == '{"a": null, "b": [1, null], "c": {"d": "ñ"}}'


def test_build_tableau_acepta_un_directorio_de_salida(tmp_path: Path) -> None:
    from scripts import build_tableau as g

    # Sin `--out` el destino se resuelve al ejecutar desde `data.tableau`; el parser no lo
    # fija, para no depender de la configuración cargada en el proceso que lo importa.
    assert g._parse_args([]).out is None
    assert g._parse_args(["--out", str(tmp_path / "staging")]).out == tmp_path / "staging"


def test_validacion_semanal_actualiza_una_copia_y_no_el_excel_real(
    tmp_path: Path, monkeypatch
) -> None:
    from scripts import genera_validacion_semanal as g

    real = tmp_path / "real" / "tabla_333_modelos_produccion.xlsx"
    real.parent.mkdir()
    wb = Workbook()
    ws = wb.active
    ws.append(["padecimiento", "entidad", "sexo", "realidad_sem_previa"])
    ws.append(["Alzheimer", "Nacional", "general", None])
    wb.save(real)
    monkeypatch.setattr(g, "PROD_EXCEL", real)

    staging = tmp_path / "staging"
    destino = g._destino_excel(staging)
    assert destino == staging / real.name and destino.is_file()

    comp = pd.DataFrame(
        [
            {
                "pad_norm": g._normalize("Alzheimer"),
                "entidad_norm": g._norm_prod_entidad("Nacional"),
                "sexo": "general",
                "real": 12.0,
            }
        ]
    )
    g._update_excel(comp, 2026, 31, destino)

    assert load_workbook(destino).active["D2"].value == 12
    assert load_workbook(real).active["D2"].value is None, "el Excel del árbol real no se toca"
    # Sin `--out` (o apuntando al directorio canónico) se conserva el comportamiento legacy.
    assert g._destino_excel(real.parent) == real
    assert g._parse_args([]).out == g.OUTPUT.parent
